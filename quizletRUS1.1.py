import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from random import randint
import datetime
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from locators import *
from main import your_mail_quizlet, your_password_quizlet

URL_MODULE = {
    '<Worksheet "Evgeniya">': 'https://quizlet.com/class/23176946/',
    '<Worksheet "Oleksii">': 'https://quizlet.com/class/23176408/',
    '<Worksheet "Valentyna_EWM">': 'https://quizlet.com/class/23129431/',
    '<Worksheet "Liubov">': 'https://quizlet.com/class/22928359/',
    '<Worksheet "Natalie">': 'https://quizlet.com/class/22928329/',
    '<Worksheet "Roman">': 'https://quizlet.com/class/22928232/',
    '<Worksheet "Valentyna">': 'https://quizlet.com/class/22882687/',
    '<Worksheet "Alexandra">': 'https://quizlet.com/class/22882663/',
    '<Worksheet "Dasha">': 'https://quizlet.com/class/22802447/',
    '<Worksheet "Polly">': 'https://quizlet.com/class/23176946/',
    '<Worksheet "Dasha_EWM">': 'https://quizlet.com/class/22802447/',
    '<Worksheet "Alex_and_Tatiana">': 'https://quizlet.com/class/22802359/',
}


class Quizlet:
    def __init__(self):

        options = webdriver.ChromeOptions()

        options.add_experimental_option("detach", True)
        self.driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
        self.driver.hold_browser_open = True
        self.driver.set_window_size(1080, 768)
        self.driver.implicitly_wait(5)

    def open_url(self, url):
        self.driver.get(url)
        return self

    def click_on_element(self, by, element):
        wait = WebDriverWait(self.driver, 10)
        by = by.upper()
        if by == 'XPATH':
            wait.until(EC.visibility_of_element_located((By.XPATH, element))).click()
        if by == 'CSS':
            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, element))).click()
        if by == 'ID':
            wait.until(EC.visibility_of_element_located((By.ID, element))).click()
        if by == 'CLASS_NAME':
            wait.until(EC.visibility_of_element_located((By.CLASS_NAME, element))).click()
        if by == 'LINK_TEXT':
            wait.until(EC.visibility_of_element_located((By.LINK_TEXT, element))).click()

    def send_keys_to_element_css(self, element, send_keys: str):
        self.driver.find_element(By.CSS_SELECTOR, element).send_keys(send_keys)
        return self

    def find_elemets(self, blok_elements):
        elements = self.driver.find_elements(By.XPATH, blok_elements)
        return elements


    def find_elemet(self, by, element):
        wait = WebDriverWait(self.driver, 10)
        by = by.upper()
        if by == 'XPATH':
            element = wait.until(EC.visibility_of_element_located((By.XPATH, element)))
        if by == 'CSS':
            element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, element)))
        if by == 'ID':
            element = wait.until(EC.visibility_of_element_located((By.ID, element)))
        if by == 'CLASS_NAME':
            element = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, element)))
        if by == 'LINK_TEXT':
            element = wait.until(EC.visibility_of_element_located((By.LINK_TEXT, element)))
        return element

    def find_elemet_css_wait20sec(self, element):
        wait = WebDriverWait(self.driver, 20)
        element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, element)))
        return element

    def delete_element(self, element_css: str):
        element = self.driver.find_element(By.CSS_SELECTOR, element_css)
        self.driver.execute_script(""" var element = arguments[0]; element.parentNode.removeChild(element); """,
                                   element)
        return self

    def find_text_element_css(self, css_locator):
        a = self.driver.find_element(By.CSS_SELECTOR, css_locator)
        return a.text

    def move_to_element(self, by, locator):
        self.driver.execute_script("return arguments[0].scrollIntoView(true);", self.find_elemet(by, locator))
        # self.driver.execute_script("arguments[0].scrollIntoView(true);", self.find_elemet(by, locator))
        # ActionChains(self.driver).scroll_to_element(self.find_elemet(by, locator)).perform()

    def choose_words_in_bloks(self):
        wait = WebDriverWait(self.driver, 5)
        bloks = self.driver.find_elements(By.XPATH, BLOKS_WORD)
        for i in bloks:
            time.sleep(1)
            i.find_element(By.XPATH, CLICK_ON_TEXT_NEXT).click()
            time.sleep(1)
            try:
                element3 = self.driver.find_element(By.XPATH, './/div[@class="h187ohc0 aixuzoy"]')
                self.driver.execute_script(""" var element = arguments[0]; element.parentNode.removeChild(element); """,
                                       element3)
            except:
                pass
            try:
                wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, SHOISE_WORD))).click()
            except:
                pass
            i.find_element(By.XPATH, IMJ_BUTTON).click()
            try:
                wait.until(EC.visibility_of_element_located(
                    (By.XPATH, f'//div[@class="ImageCarousel-imagePage"][1]//div[{randint(1, 4)}]'))).click()
            except:
                pass


    def push_ctrl_with_any(self, botm: str):
        ActionChains(self.driver).key_down(Keys.CONTROL).send_keys(botm).key_up(Keys.CONTROL).perform()

    def new_quizlet_open(self, url: str, ):
        self.driver.execute_script(f"window.open ('{url}')")

    def switch_window(self, num):
        self.driver.switch_to.window(self.driver.window_handles[num])


    def open_exel(self):
        self.wb = openpyxl.load_workbook('people.xlsx')

    def sheets_number(self):
        sheet_number = len(self.wb.sheetnames)
        return sheet_number

    def name_exel_sheet(self, sheet_num: int):
        self.wb.active = sheet_num
        return self.wb.active

    def reed_exel_cell(self, cell: str):
        exel_cell = self.wb.active[cell].value
        return exel_cell

    def number_of_lines(self):
        sheets_list = self.wb.sheetnames
        sheet_active = self.wb[sheets_list[0]]
        row_max = sheet_active.max_row
        return row_max


def main():
    quizlet = Quizlet()
    quizlet.open_exel()
    quizlet.open_url('https://quizlet.com/ru')
    quizlet.switch_window(0)
    try:
        quizlet.click_on_element('CSS', BUTTON_ACCEPT_COOKIE)
    except:
        pass
    quizlet.click_on_element('CSS', BUTTON_LOGIN_1)
    quizlet.send_keys_to_element_css('#username', your_mail_quizlet)
    quizlet.send_keys_to_element_css('#password', your_password_quizlet)
    quizlet.move_to_element('CSS', BUTTON_LOGIN_2)
    time.sleep(1)
    quizlet.click_on_element('CSS', BUTTON_LOGIN_2)
    quizlet.find_elemet_css_wait20sec(LOCATOR_AFTER_AUTORIZATION)
    num = 1
    a = 1
    for i in range(quizlet.sheets_number()):
        quizlet.open_url(URL_MODULE.get(f'{quizlet.name_exel_sheet(i)}'))
        time.sleep(0.5)
        text_module_name = quizlet.find_text_element_css(MODULE_NAME)
        quizlet.click_on_element('CSS', PRESS_PLUS)
        quizlet.click_on_element('CSS', CREATE_NEW_MODULE)
        if i < 1:
            quizlet.click_on_element('CSS', CLOSE_AD)
        quizlet.click_on_element('CSS', PRESS_IMPORT_WORLD)
        for w in range(1, 999):
            if quizlet.reed_exel_cell("A" + str(w)) is None:
                break
            quizlet.send_keys_to_element_css('.ImportTerms-textarea',
                                                str(f'{quizlet.reed_exel_cell("A" + str(w))}\n'))

        time.sleep(2)
        quizlet.click_on_element('CSS', PRESS_IMPORT_BUTTOM)
        quizlet.send_keys_to_element_css(MODULE_NAME_TYPE,
                                         f'{text_module_name} {datetime.datetime.today().strftime("%d.%m")}')
        time.sleep(1)
        quizlet.move_to_element('CSS', '.TermRows [class="TermRows-termRowWrap"]:first-child')
        quizlet.click_on_element('XPATH', CLICK_ON_TEXT)
        quizlet.click_on_element('XPATH', PRESS_LANGUAGE_SHOISE1)
        quizlet.click_on_element('XPATH', SHOISE_LANGUAGE_ENG)
        quizlet.click_on_element('XPATH', PRESS_LANGUAGE_SHOISE2)
        quizlet.click_on_element('XPATH', SHOISE_LANGUAGE_RUS)
        quizlet.delete_element('[placeholder="Поиск языков"]')
        quizlet.delete_element('[class="LanguageSelect-input"]')
        quizlet.choose_words_in_bloks()
        time.sleep(1)
        if a < quizlet.sheets_number() :
            quizlet.new_quizlet_open('https://quizlet.com/inna_papanova/classes')
            quizlet.switch_window(num)
            num += 1
            a +=1




if __name__ == '__main__':
    main()
